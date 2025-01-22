import logging
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackContext
from reportlab.lib.pagesizes import A5, landscape
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
import openpyxl
import re
from io import BytesIO
import os

# Включаем логирование
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)

# Вставьте сюда свой API токен
API_TOKEN = 'API_KEY'

# Регистрируем шрифты
pdfmetrics.registerFont(TTFont('DejaVuSans', 'fonts/DejaVuSans.ttf'))
pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', 'fonts/DejaVuSans-Bold.ttf'))

def format_value(value):
    if isinstance(value, (int, float)):
        return f"{value:.0f}" if value == int(value) else f"{value:.2f}"
    elif isinstance(value, str):
        return value.strip()
    return value if value else "-"

def format_date(value):
    if isinstance(value, str):
        return value.split(" ")[0]
    elif hasattr(value, 'strftime'):
        return value.strftime("%d.%m.%Y")
    return "-"

def create_payslips(excel_file, month_year):
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb[month_year]

    payslips = []
    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
        employee_name = row[0].value
        if not employee_name or row[0].row > 250:
            continue
        if row[0].font.color and row[0].font.color.rgb == 'FFFF0000':
            continue
        if re.match(r'^[А-ЯЁ][а-яё]+ [А-ЯЁ]\.[А-ЯЁ]\.$', str(employee_name)):
            continue

        total_hours = sum(
            [row[37].value or 0, row[38].value or 0, row[35].value or 0, row[36].value or 0]
        )
        shifts = row[37].value or 0
        overtime = row[38].value or 0
        vacation = row[35].value or 0
        sick_leave = row[36].value or 0
        rank = row[34].value or 0

        summary_data = [
            ["Всего (ч)", total_hours],
            ["Смены (ч)", shifts],
            ["Переработки (ч)", overtime],
            ["Отпуск (ч)", vacation],
            ["Больничный (ч)", sick_leave],
            ["Разряд", rank]
        ]

        payslip = {
            'Сотрудник': format_value(employee_name),
            'Должность': format_value(row[2].value),
            'Дата трудоустройства': format_date(row[6].value),
            'Месяцев отработано': format_value(row[7].value),
            'Количество отработанных смен': format_value(row[10].value),
            'График работы': format_value(row[9].value),
            'Часов (день)': format_value(row[12].value),
            'Часов (ночь)': format_value(row[13].value),
            'Оклад (день)': format_value(row[14].value),
            'Налог': format_value(row[16].value) if row[16].value is not None else "0",
            'Больничный': format_value(row[17].value),
            'Отпуск': format_value(row[18].value),
            'Надбавка за стаж': format_value(row[20].value),
            'Доплаты за умения': format_value(row[22].value),
            'Переработки (часы день)': format_value(row[23].value),
            'Переработки (часы ночь)': format_value(row[24].value),
            'Переработки (оплата)': format_value(row[25].value),
            'Премия руководства': format_value(row[29].value),
            'Депремирование': format_value(row[30].value),
            'Итого': format_value(row[31].value),
            'Примечание': format_value(row[3].value),
        }

        payslip = {k: v for k, v in payslip.items() if v not in ["-", "0"] or k == "Налог"}
        payslips.append((payslip, summary_data))

    pdf_filename = f"payslips_{month_year}.pdf"
    c = canvas.Canvas(pdf_filename, pagesize=landscape(A5))

    for payslip, summary_data in payslips:
        c.setFont("DejaVuSans-Bold", 14)
        c.setFillColor(colors.grey)
        c.drawString(20, 390, f"Расчетный лист: {month_year}")

        c.drawImage("logo1.jpg", 480, 300, width=100, height=100)

        data = [[key, value] for key, value in payslip.items() if key not in ["Итого", "Примечание"]]
        table = Table(data, colWidths=[180, 220])

        table_style = TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
            ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ])

        for i, (key, value) in enumerate(payslip.items()):
            if key == 'Налог':
                table_style.add('TEXTCOLOR', (1, i), (1, i), colors.red)
            elif key in ['Оклад (день)', 'Больничный', 'Отпуск', 'Надбавка за стаж', 'Доплаты за умения', 'Итого']:
                table_style.add('TEXTCOLOR', (1, i), (1, i), colors.green)
            elif key == 'Депремирование':
                table_style.add('TEXTCOLOR', (1, i), (1, i), colors.red)
            elif key == 'Количество отработанных смен':
                table_style.add('FONTSIZE', (0, i), (0, i), 9)
            elif key == 'Должность' and len(value) > 35 and len(value) <= 42:
                table_style.add('FONTSIZE', (1, i), (1, i), 9)
            elif key == 'Должность' and len(value) > 42:
                table_style.add('FONTSIZE', (1, i), (1, i), 7)

        table.setStyle(table_style)

        table_width, table_height = table.wrap(0, 0)
        table_top_y = 370 - table_height
        table.drawOn(c, 20, table_top_y)

        # Добавляем таблицу справа по центру от основной
        summary_table = Table(summary_data, colWidths=[110, 40])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
            ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        summary_table.wrapOn(c, 0, 0)
        summary_table.drawOn(c, 430, 170)

        c.setFillColor(colors.red)
        c.rect(460, 90, 120, 40, fill=1)

        c.setFillColor(colors.white)
        c.setFont("DejaVuSans-Bold", 12)
        c.drawString(465, 105, "Аванс: 25000")

        total_value = payslip.get('Итого', '-')
        c.setFillColor(colors.green)
        c.rect(460, 30, 120, 40, fill=1)

        c.setFillColor(colors.white)
        c.setFont("DejaVuSans-Bold", 12)
        c.drawString(465, 45, f"Итого: {total_value}")

        note = payslip.get('Примечание', '')
        if note:
            c.setFont("DejaVuSans", 10)
            c.setFillColor(colors.black)
            c.drawString(20, 20, f"Примечание: {note}")

        c.showPage()

    c.save()


# Функция для старта
async def start(update: Update, context: CallbackContext) -> None:
    await update.message.reply_text('Привет! Отправьте мне Excel файл и укажите месяц и год для генерации PDF.')

# Функция обработки получения файла
async def handle_document(update: Update, context: CallbackContext) -> None:
    document = update.message.document
    if document.mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        file = await document.get_file()
        # Скачиваем файл как байтовый массив
        file_data = await file.download_as_bytearray()
        
        # Сохраняем байтовый массив в файл
        with open('uploaded_file.xlsx', 'wb') as f:
            f.write(file_data)
        
        await update.message.reply_text('Файл получен. Укажите месяц и год (например, "Сентябрь 2024").')
    else:
        await update.message.reply_text('Пожалуйста, отправьте файл формата Excel (.xlsx).')

# Функция для обработки текста (месяц и год)
async def handle_text(update: Update, context: CallbackContext) -> None:
    month_year = update.message.text.strip()
    if os.path.exists('uploaded_file.xlsx'):
        try:
            # Генерируем PDF с расчетными листами
            create_payslips('uploaded_file.xlsx', month_year)
            
            # Отправляем PDF файл обратно в чат
            with open(f"payslips_{month_year}.pdf", 'rb') as f:
                await update.message.reply_document(f)
            
            os.remove(f"payslips_{month_year}.pdf")
        except Exception as e:
            await update.message.reply_text(f"Ошибка: {e}")
    else:
        await update.message.reply_text('Пожалуйста, сначала отправьте Excel файл.')

def main() -> None:
    application = Application.builder().token(API_TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    application.run_polling()

if __name__ == '__main__':
    main()